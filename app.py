from flask import Flask, render_template, request, redirect, send_file, url_for
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import pandas as pd
from werkzeug.utils import secure_filename

app = Flask(__name__)

EXCEL_FILE = "mental_log.xlsx"
UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

SHEETS = {
    "DailyLog": [
    "วันที่เวลา", "Mood", "Anxiety", "Depress",
    "Happy", "Irritable", "Trigger", "Note"
    ],
    "FlashbackLog": [
        "วันที่เวลา", "กิจกรรมที่กำลังทำ", "ภาพ/ความทรงจำที่ผุดขึ้น",
        "Trigger", "อาการที่เกิดขึ้น", "ความรุนแรง 0-10",
        "วิธี Cope", "หลัง Cope เหลือกี่คะแนน"
    ],
    "RuminationLog": [
        "วันที่เวลา", "Topic", "ความคิดวน", "อารมณ์ที่เกิด",
        "คำตอบแบบเมตตาตัวเอง"
    ],
    "CopingNotes": [
        "วันที่เวลา", "หมวด", "ชื่อ Note", "ข้อความ"
    ],
    "HopeBox": [
        "วันที่เวลา", "ประเภท", "ชื่อ", "รายละเอียด", "รูป/ลิงก์"
    ],
    "Assessments": [
        "วันที่เวลา", "แบบประเมิน", "คะแนนรวม", "แปลผล"
    ],
    "HealingLog": [
    "วันที่เวลา", "คำตอบทั้งหมด", "สรุปข้อความฮีลใจ"
    ],
}


def init_excel():
    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()
    first = True

    for sheet_name, headers in SHEETS.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        ws.append(headers)
        first = False

    wb.save(EXCEL_FILE)


def append_excel(sheet_name, row):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    ws.append(row)
    wb.save(EXCEL_FILE)


def read_sheet(sheet_name):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    rows = list(ws.values)

    if len(rows) <= 1:
        return []

    headers = rows[0]
    data = []

    for row in rows[1:]:
        item = {}
        for i, header in enumerate(headers):
            item[header] = row[i] if i < len(row) else ""
        data.append(item)

    return list(reversed(data))


def interpret_gad7(score):
    if score <= 5:
        return "วิตกกังวลเล็กน้อย"
    elif score <= 10:
        return "วิตกกังวลปานกลาง"
    elif score <= 15:
        return "วิตกกังวลสูง"
    return "วิตกกังวลสูงมาก"


def interpret_phq9(score):
    if score <= 4:
        return "น้อยมาก"
    elif score <= 9:
        return "เล็กน้อย"
    elif score <= 14:
        return "ปานกลาง"
    elif score <= 19:
        return "ค่อนข้างรุนแรง"
    return "รุนแรง"


def interpret_pclt(score):
    if score <= 29:
        return "อาการ PTSD น้อย"
    elif score <= 44:
        return "อาการ PTSD ปานกลาง"
    return "อาการ PTSD สูง"


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        append_excel("DailyLog", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            request.form.get("mood"),
            request.form.get("anxiety"),
            request.form.get("depress"),
            request.form.get("happy"),
            request.form.get("irritable"),
            request.form.get("trigger"),
            request.form.get("note"),
        ])
        return redirect(url_for("index"))

    logs = read_sheet("DailyLog")[:7]
    latest = logs[0] if logs else None
    return render_template("index.html", logs=logs, latest=latest)


@app.route("/flashback", methods=["GET", "POST"])
def flashback():
    if request.method == "POST":
        append_excel("FlashbackLog", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            request.form.get("activity"),
            request.form.get("flashback_image"),
            request.form.get("trigger"),
            request.form.get("symptoms"),
            request.form.get("intensity"),
            request.form.get("coping"),
            request.form.get("after_score"),
        ])
        return redirect(url_for("flashback"))

    items = read_sheet("FlashbackLog")
    return render_template("flashback.html", items=items)


@app.route("/rumination", methods=["GET", "POST"])
def rumination():
    if request.method == "POST":
        append_excel("RuminationLog", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            request.form.get("topic"),
            request.form.get("thought"),
            request.form.get("emotion"),
            request.form.get("reply"),
        ])
        return redirect(url_for("rumination"))

    items = read_sheet("RuminationLog")
    return render_template("rumination.html", items=items)


@app.route("/coping", methods=["GET", "POST"])
def coping():
    if request.method == "POST":
        append_excel("CopingNotes", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            request.form.get("category"),
            request.form.get("title"),
            request.form.get("content"),
        ])
        return redirect(url_for("coping"))

    items = read_sheet("CopingNotes")
    return render_template("coping.html", items=items)


@app.route("/hopebox", methods=["GET", "POST"])
def hopebox():
    if request.method == "POST":
        file_path = request.form.get("link") or ""

        file = request.files.get("image")
        if file and file.filename:
            filename = datetime.now().strftime("%Y%m%d%H%M%S_") + secure_filename(file.filename)
            save_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(save_path)
            file_path = "/" + save_path.replace("\\", "/")

        append_excel("HopeBox", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            request.form.get("type"),
            request.form.get("title"),
            request.form.get("detail"),
            file_path,
        ])
        return redirect(url_for("hopebox"))

    items = read_sheet("HopeBox")
    return render_template("hopebox.html", items=items)


@app.route("/report")
def report():
    daily = read_sheet("DailyLog")
    flashbacks = read_sheet("FlashbackLog")
    ruminations = read_sheet("RuminationLog")
    assessments = read_sheet("Assessments")
    coping_notes = read_sheet("CopingNotes")
    hopebox = read_sheet("HopeBox")

    def avg(items, key):
        values = []
        for item in items:
            try:
                if item.get(key) not in [None, ""]:
                    values.append(float(item.get(key)))
            except:
                pass
        return round(sum(values) / len(values), 1) if values else "-"

    summary = {
        "total_daily": len(daily),
        "total_flashback": len(flashbacks),
        "total_rumination": len(ruminations),
        "avg_mood": avg(daily, "Mood"),
        "avg_anxiety": avg(daily, "Anxiety"),
        "avg_depress": avg(daily, "Depress"),
        "avg_happy": avg(daily, "Happy"),
        "avg_irritable": avg(daily, "Irritable"),
    }

    return render_template(
        "report.html",
        daily=daily,
        flashbacks=flashbacks,
        ruminations=ruminations,
        assessments=assessments,
        coping_notes=coping_notes,
        hopebox=hopebox,
        summary=summary
    )


@app.route("/export")
def export_excel():
    init_excel()
    return send_file(EXCEL_FILE, as_attachment=True)

@app.route("/doctor-pdf")
def doctor_pdf():
    daily = read_sheet("DailyLog")
    flashbacks = read_sheet("FlashbackLog")
    ruminations = read_sheet("RuminationLog")
    assessments = read_sheet("Assessments")

    def avg(items, key):
        values = []
        for item in items:
            try:
                v = item.get(key)
                if v not in [None, ""]:
                    values.append(float(v))
            except:
                pass
        return round(sum(values) / len(values), 1) if values else "-"
    
    def latest_assessment_score(items, assessment_type):
        for item in items:
            if item.get("แบบประเมิน") == assessment_type:
                return {
                    "score": item.get("คะแนนรวม") or "-",
                    "result": item.get("แปลผล") or "-",
                    "date": item.get("วันที่เวลา") or "-"
                }
        return {
            "score": "-",
            "result": "-",
            "date": "-"
        }

    summary = {
        "total_daily": len(daily),
        "total_flashback": len(flashbacks),
        "total_rumination": len(ruminations),
        "avg_mood": avg(daily, "Mood"),
        "avg_anxiety": avg(daily, "Anxiety"),
        "avg_depress": avg(daily, "Depress"),
        "avg_happy": avg(daily, "Happy"),
        "avg_irritable": avg(daily, "Irritable"),
        "gad7": latest_assessment_score(assessments, "GAD-7"),
        "phq9": latest_assessment_score(assessments, "PHQ-9"),
        "pclt": latest_assessment_score(assessments, "PCL-T"),
    }

    return render_template(
        "doctor_pdf.html",
        daily=daily,
        flashbacks=flashbacks,
        ruminations=ruminations,
        assessments=assessments,
        summary=summary
    )


@app.route("/assessments", methods=["GET", "POST"])
def assessments():
    questions = {
        "GAD-7": [
            "รู้สึกกระวนกระวาย กังวล หรือกระสับกระส่าย",
            "ไม่สามารถหยุดหรือควบคุมความกังวลได้",
            "กังวลมากเกินไปในหลายๆ เรื่อง",
            "มีปัญหาเรื่องการผ่อนคลายอารมณ์",
            "กระสับกระส่ายจนไม่สามารถอยู่นิ่งๆ ได้",
            "กลายเป็นคนขี้รำคาญ หรือหงุดหงิดง่าย",
            "รู้สึกกลัวว่าจะมีเหตุการณ์ร้ายๆ เกิดขึ้น",
        ],
        "PHQ-9": [
            "ไม่ค่อยอยากทำ หรือไม่รู้สึกสนุกที่จะทำอะไร",
            "รู้สึกเศร้า หดหู่ หรือสิ้นหวัง",
            "มีปัญหานอนไม่หลับ หลับไม่สนิท หรือนอนมากเกินไป",
            "รู้สึกเหนื่อยหรือไม่ค่อยมีแรง",
            "ไม่ค่อยอยากกินอาหาร หรือกินมากเกินไป",
            "รู้สึกแย่กับตัวเอง หรือคิดว่าตัวเองล้มเหลว",
            "ไม่มีสมาธิในการทำสิ่งต่างๆ",
            "เคลื่อนไหวหรือพูดช้าลง หรือกระสับกระส่ายมากกว่าปกติ",
            "คิดว่าตายไปเสียได้ก็ดี หรือคิดทำร้ายตัวเอง",
        ],
        "PCL-T": [
            "มีสิ่งต่างๆ เกี่ยวกับเหตุการณ์ผุดขึ้นมาซ้ำๆ เช่น ภาพ ความคิด หรือการรับรู้",
            "มีความฝันที่ทำให้ทุกข์ทรมานซ้ำๆ เกี่ยวกับเหตุการณ์",
            "รู้สึกเหมือนเหตุการณ์นั้นเกิดขึ้นมาอีกครั้ง",
            "มีความทุกข์ใจมากเมื่อเจอสิ่งที่ทำให้ระลึกถึง",
            "มีปฏิกิริยาทางร่างกาย เช่น ใจสั่น หายใจไม่อิ่ม เหงื่อแตก เมื่อระลึกถึง",
            "พยายามเลี่ยงความคิด ความรู้สึก หรือการสนทนาที่เกี่ยวข้อง",
            "พยายามเลี่ยงกิจกรรม สถานที่ หรือบุคคลที่กระตุ้นให้นึกถึง",
            "ไม่สามารถระลึกถึงส่วนสำคัญของเหตุการณ์นั้น",
            "ความสนใจหรือการเข้าร่วมกิจกรรมสำคัญลดลง",
            "รู้สึกแปลกแยกหรือห่างเหินจากผู้อื่น",
            "ขอบเขตอารมณ์ลดลง รู้สึกชาหรือรู้สึกอะไรได้น้อยลง",
            "มองอนาคตไม่ยาวไกล",
            "นอนหลับยากหรือหลับๆ ตื่นๆ",
            "หงุดหงิดหรือโกรธง่าย",
            "ตั้งสมาธิลำบาก",
            "ระวังระไวมากกว่าปกติ",
            "สะดุ้งตกใจมากกว่าปกติ",
        ],
    }

    if request.method == "POST":
        assessment_type = request.form.get("assessment_type")
        scores = []

        for i in range(1, len(questions[assessment_type]) + 1):
            scores.append(int(request.form.get(f"q{i}", 0)))

        total = sum(scores)

        if assessment_type == "GAD-7":
            result = interpret_gad7(total)
        elif assessment_type == "PHQ-9":
            result = interpret_phq9(total)
        else:
            result = interpret_pclt(total)

        append_excel("Assessments", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            assessment_type,
            total,
            result,
        ])

        return redirect(url_for("assessments"))

    items = read_sheet("Assessments")
    return render_template("assessments.html", questions=questions, items=items)


HEALING_QUESTION_SETS = {
    "flashback": {
        "title": "Flashback Grounding",
        "subtitle": "พาตัวเองกลับมาสู่ปัจจุบันอย่างช้าๆ",
        "questions": [
            "ตอนนี้มีอะไรบอกเราได้บ้างว่าเราอยู่ในปัจจุบัน ไม่ใช่ตอนนั้น",
            "ร่างกายส่วนไหนกำลังเกร็งหรือกลัวที่สุด",
            "สิ่งเล็กๆ ที่ช่วยให้รู้สึกปลอดภัยขึ้น 1% คืออะไร",
            "มีอะไรในห้องนี้ที่จับต้องได้จริง 3 อย่าง",
            "ตอนนี้เราต้องการ grounding แบบไหน",
            "ถ้าความกลัวนี้พูดได้ มันกำลังพยายามปกป้องเราจากอะไร",
            "ประโยคสั้นๆ ที่อยากบอกตัวเองตอน flashback คืออะไร"
        ]
    },
    "guilt": {
        "title": "Self Compassion",
        "subtitle": "ลดการโทษตัวเอง แล้วค่อยๆ กลับมาพูดกับตัวเองอย่างอ่อนโยน",
        "questions": [
            "เรื่องที่เรากำลังโทษตัวเองคืออะไร",
            "ถ้าเพื่อนรักเจอเรื่องเดียวกัน เราจะพูดกับเขาว่าอะไร",
            "ส่วนไหนของเรื่องนี้ที่เราไม่ได้ตั้งใจให้เกิดขึ้น",
            "เราแบกรับความรับผิดชอบเกินจริงตรงไหน",
            "สิ่งที่เราอยากให้อภัยตัวเอง แม้ยังทำไม่ได้เต็มที่ คืออะไร",
            "วันนี้เราทำดีที่สุดภายใต้เงื่อนไขอะไรบ้าง",
            "ประโยคเมตตาที่อยากฝากไว้ให้ตัวเองคืออะไร"
        ]
    },
    "overwhelmed": {
        "title": "Overwhelmed Reset",
        "subtitle": "วางภาระลงทีละนิด เมื่อทุกอย่างเยอะเกินไป",
        "questions": [
            "ตอนนี้อะไรเยอะเกินไปสำหรับใจเรา",
            "เรื่องไหนที่ยังไม่ต้องแก้วันนี้ก็ได้",
            "ถ้าลดภาระลง 10% เราจะวางอะไรลงก่อน",
            "ร่างกายกำลังขอพักแบบไหน",
            "หนึ่งอย่างเล็กๆ ที่พอทำได้ใน 5 นาทีคืออะไร",
            "ใครหรืออะไรที่ช่วยให้เราไม่ต้องอยู่คนเดียวกับเรื่องนี้",
            "ประโยคที่อยากบอกตัวเองตอนหมดแรงคืออะไร"
        ]
    },
    "night": {
        "title": "Night Anxiety",
        "subtitle": "สำหรับตอนกลางคืนที่ใจคิดวนและรู้สึกไม่ปลอดภัย",
        "questions": [
            "ความกลัวตอนกลางคืนนี้กำลังบอกอะไรเรา",
            "ตอนนี้มีหลักฐานอะไรบ้างว่าเรายังปลอดภัยพอ",
            "ร่างกายกำลังส่งสัญญาณอะไร และเราแปลมันว่าอะไร",
            "ถ้าความคิดนี้เป็นแค่คลื่นหนึ่งลูก เราจะปล่อยให้มันผ่านไปยังไง",
            "อะไรช่วยให้คืนนี้เบาลงได้ 1%",
            "พรุ่งนี้ค่อยคิดเรื่องอะไรแทนคืนนี้ได้บ้าง",
            "ประโยคสั้นๆ ที่อยากอ่านก่อนนอนคืออะไร"
        ]
    },
    "recovery": {
        "title": "Gentle Recovery",
        "subtitle": "สำหรับวันที่หมดแรง เฉยชา หรือไม่อยากทำอะไร",
        "questions": [
            "วันนี้พลังงานของเราเหลือประมาณไหน",
            "อะไรที่ทำให้เราหมดแรงที่สุด",
            "สิ่งเล็กที่สุดที่ยังพอทำได้โดยไม่ฝืนคืออะไร",
            "ร่างกายอยากได้รับการดูแลแบบไหน",
            "เราผ่านอะไรมาแล้วบ้าง แม้มันจะดูเล็กมาก",
            "ถ้าวันนี้ไม่ต้องเก่ง เราอยากอนุญาตให้ตัวเองทำอะไร",
            "ประโยคที่อยากฝากไว้ให้ตัวเองในวันที่อ่อนล้าคืออะไร"
        ]
    }
}


@app.route("/healing", methods=["GET", "POST"])
def healing():
    theme = request.args.get("theme", "flashback")
    selected = HEALING_QUESTION_SETS.get(theme, HEALING_QUESTION_SETS["flashback"])
    questions = selected["questions"]

    if request.method == "POST":
        theme = request.form.get("theme", "flashback")
        selected = HEALING_QUESTION_SETS.get(theme, HEALING_QUESTION_SETS["flashback"])
        questions = selected["questions"]

        answers = []
        for i in range(1, len(questions) + 1):
            answers.append(request.form.get(f"answer{i}", ""))

        summary_text = f"""
            Healing Theme: {selected["title"]}

            วันนี้ใจของเรากำลังอยู่กับเรื่อง:
            "{answers[1] if len(answers) > 1 else ''}"

            สิ่งที่ยังพอช่วยพยุงเราได้ตอนนี้คือ:
            "{answers[2] if len(answers) > 2 else ''}"

            สิ่งที่ร่างกายหรือใจต้องการคือ:
            "{answers[4] if len(answers) > 4 else ''}"

            ประโยคที่อยากฝากไว้ให้ตัวเองคือ:
            "{answers[-1] if answers else ''}"

            ค่อยๆ อยู่กับตอนนี้ ไม่ต้องรีบหาย ไม่ต้องเก่งตลอดเวลา แค่กลับมาอยู่กับตัวเองทีละนิดก็พอแล้ว
            """

        all_answers = "\n".join([
            f"{idx + 1}. {questions[idx]}: {answers[idx]}"
            for idx in range(len(questions))
        ])

        append_excel("HealingLog", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            all_answers,
            summary_text.strip()
        ])

        append_excel("CopingNotes", [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            f"Healing Journey - {selected['title']}",
            "ข้อความจาก Healing Journey",
            summary_text.strip()
        ])

        return render_template(
            "healing_result.html",
            summary_text=summary_text.strip()
        )

    return render_template(
        "healing.html",
        questions=questions,
        theme=theme,
        selected=selected,
        question_sets=HEALING_QUESTION_SETS


        
    )


def ensure_all_sheets():
    init_excel()
    wb = load_workbook(EXCEL_FILE)

    for sheet_name, headers in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)

    wb.save(EXCEL_FILE)


@app.route("/bubble")
def bubble():
    comfort_lines = [

    "ตอนนี้เราไม่จำเป็นต้องเก่งก็ได้",
    "ค่อยๆ หายใจช้าๆ ทีละรอบก็พอ",
    "ความคิดไม่ใช่อันตรายเสมอไป",
    "ตอนนี้ปลอดภัยพอแล้ว",
    "เราไม่ต้องรีบหายก็ได้",
    "คืนนี้แค่ประคองตัวเองก็เก่งมากแล้ว",
    "ใจที่เหนื่อย ไม่ได้แปลว่าอ่อนแอ",
    "แค่ยังอยู่ตรงนี้ก็เพียงพอแล้ว",
    "เราผ่านวันที่ยากกว่านี้มาได้แล้ว",
    "ร่างกายที่กลัว กำลังพยายามปกป้องเรา",
    "ตอนนี้ไม่จำเป็นต้องแก้ทุกอย่าง",
    "ค่อยๆ อยู่กับปัจจุบันทีละนิด",
    "เราไม่จำเป็นต้องแบกทุกอย่างคนเดียว",
    "พักได้ ไม่ผิดเลย",
    "เราไม่ได้ล้มเหลว แค่เหนื่อยมาก",
    "ไม่เป็นไรถ้าวันนี้ทำได้ไม่มาก",
    "วันนี้รอดมาก็ถือว่าเก่งแล้ว",
    "เรายังมีเวลาค่อยๆ ดีขึ้น",
    "ความรู้สึกนี้จะไม่อยู่ตลอดไป",
    "คืนนี้ยังไม่ต้องหาคำตอบทุกเรื่อง",
    "เราไม่ต้องพิสูจน์คุณค่าของตัวเองตอนนี้",
    "ความกลัวไม่ใช่ข้อเท็จจริงเสมอไป",
    "ตอนนี้ยังไม่ต้องตัดสินตัวเอง",
    "แค่หายใจต่อก็พอแล้ว",
    "ค่อยๆ ผ่อนไหล่ลงนะ",
    "เรายังสมควรได้รับความอ่อนโยน",
    "สิ่งที่เกิดขึ้นกับเรา มันหนักจริง",
    "เราไม่ได้อ่อนแอที่ยังเจ็บอยู่",
    "วันนี้ไม่ต้อง productive ก็ได้",
    "เราไม่จำเป็นต้องโอเคทันที",
    "ตอนนี้หัวใจเราคงเหนื่อยมากเลย",
    "ไม่เป็นไรถ้ายัง move on ไม่ได้",
    "ค่อยๆ กลับมาอยู่กับร่างกายตัวเองนะ",
    "ตอนนี้เราปลอดภัยกว่าที่ใจคิด",
    "ทุกอย่างไม่ต้องดีพร้อมในคืนนี้",
    "เราไม่ต้องรีบตอบทุกคำถามในหัว",
    "บางวันเป้าหมายคือแค่ผ่านวันไปให้ได้",
    "เราไม่ได้อยู่คนเดียวกับความรู้สึกนี้",
    "ค่อยๆ วางเรื่องที่หนักลงทีละนิด",
    "ไม่เป็นไรถ้ายังสับสนอยู่",
    "ตอนนี้เราไม่ต้องเข้มแข็งตลอดเวลาก็ได้",
    "ร่างกายเรากำลังเหนื่อยจากการระวังตัวมาตลอด",
    "เราอนุญาตให้ตัวเองพักได้",
    "ตอนนี้ยังไม่ต้องคิดถึงอนาคตทั้งหมด",
    "ใจเรากำลังพยายามดีที่สุดแล้ว",
    "คืนนี้เอาแค่ให้ตัวเองปลอดภัยก่อน",
    "ไม่จำเป็นต้องหายไวเพื่อมีคุณค่า",
    "เราไม่ผิดที่ยังรู้สึกอยู่",
    "ค่อยๆ กลับมาที่ลมหายใจ",
    "วันนี้เราก็พยายามมากแล้ว",
    "ไม่มีใครควรต้องแบกความเจ็บปวดนี้คนเดียว",
    "เราไม่ต้องเก่งเพื่อสมควรถูกรัก",
    "บางอย่างต้องใช้เวลา และนั่นไม่ผิดเลย",
    "ตอนนี้พักก่อนก็ได้",
    "เราไม่ได้ถอยหลัง แค่กำลังเหนื่อย",
    "ร่างกายที่ตื่นกลัว ไม่ได้หมายความว่าอันตรายกำลังเกิดขึ้น",
    "คืนนี้ขอให้ใจเราเบาลงอีกนิด",
    "เราไม่จำเป็นต้องรับมือทุกอย่างพร้อมกัน",
    "แค่ยังอยู่ตรงนี้ ก็ถือว่าเก่งมากแล้ว",
    "เราอนุญาตให้ตัวเองค่อยๆ ฟื้นได้",
]
    return render_template(
        "bubble.html",
        comfort_lines=comfort_lines
    )

@app.route("/sand")
def sand():
    return render_template("sand.html")

@app.route("/rain")
def rain():
    return render_template("rain.html")

@app.route("/firefly")
def firefly():

    comfort_lines = [
        "คืนนี้ค่อยๆ พักนะ 🌙",
        "เราไม่จำเป็นต้องเก่งตลอดเวลา",
        "หายใจช้าๆ แล้วอยู่กับแสงเล็กๆ นี้",
        "ไม่เป็นไรถ้ายังเหนื่อยอยู่",
        "คืนนี้ยังไม่ต้องแก้ทุกเรื่อง",
        "ใจที่อ่อนล้า ควรได้รับความอ่อนโยน",
        "ค่อยๆ อยู่กับปัจจุบันทีละนิด",
        "ตอนนี้ปลอดภัยพอแล้ว",
        "เราไม่ได้อยู่คนเดียวกับความรู้สึกนี้",
        "แค่ยังอยู่ตรงนี้ก็พอแล้ว"
    ]

    return render_template(
        "firefly.html",
        comfort_lines=comfort_lines
    )

@app.route("/cloud")
def cloud():
    return render_template("cloud.html")

if __name__ == "__main__":
    init_excel()
    ensure_all_sheets()
    app.run(debug=True, host="0.0.0.0", port=5000)